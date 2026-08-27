import sys
import os
import re
from pathlib import Path
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QFileDialog, QLabel, QProgressBar, QTextEdit, QMessageBox
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QIcon
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import traceback


class DashboardGenerator(QThread):
    """Worker thread to generate dashboard"""
    progress_update = pyqtSignal(str)
    progress_complete = pyqtSignal(bool, str)
    REQUIRED_COLUMNS = [
        "BUSSINESS LINES",
        "APPLICATION",
        "SBOM",
        "APPLICATION TYPE",
        "SBOM STATUS",
        "TOTAL COMPONENTS COUNT",
        "CC AUTOMATION COUNT",
        "MANUALLY  ONBOARDED COMPONENTS",
        "TOTAL COMPONENTS ONBOARDED",
        "COMPONENTS PENDING TO BE ONBOARDED",
        "JIRA TICKET",
    ]
    NUMERIC_COLUMNS = [
        "TOTAL COMPONENTS COUNT",
        "CC AUTOMATION COUNT",
        "MANUALLY  ONBOARDED COMPONENTS",
        "TOTAL COMPONENTS ONBOARDED",
        "COMPONENTS PENDING TO BE ONBOARDED",
    ]
    
    def __init__(self, input_file):
        super().__init__()
        self.input_file = input_file
        self.output_file = None

    @staticmethod
    def _normalize(value):
        return "".join(ch.lower() for ch in str(value) if ch.isalnum())

    @staticmethod
    def _normalize_sheet_name(value):
        return "".join(ch.lower() for ch in str(value) if ch.isalnum())

    def _coerce_numeric_value(self, ws, value, depth=0):
        if depth > 6:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if value is None:
            return None

        text = str(value).strip()
        if text == "":
            return None

        if text.startswith("="):
            expr = text[1:].replace("$", "")
            refs = set(re.findall(r"[A-Za-z]{1,3}[0-9]+", expr))
            for ref in refs:
                ref_value = ws[ref].value
                ref_num = self._coerce_numeric_value(ws, ref_value, depth + 1)
                if ref_num is None:
                    ref_num = 0.0
                expr = re.sub(rf"\b{ref}\b", str(ref_num), expr, flags=re.IGNORECASE)

            if not re.fullmatch(r"[0-9eE\.\+\-\*\/\(\) ]+", expr):
                return None

            try:
                return float(eval(expr, {"__builtins__": {}}, {}))
            except Exception:
                return None

        try:
            return float(text)
        except Exception:
            return None

    def _build_summary_frames(self, combined_df):
        combined_df = combined_df.copy()
        combined_df["SBOM Completed"] = (
            combined_df["SBOM STATUS"].astype(str).str.strip().str.lower() == "completed"
        )

        sbom_tracker = combined_df[[
            "Business Segment",
            "Bussiness Lines",
            "APPLICATION",
            "SBOM",
            "SBOM STATUS",
            "TOTAL COMPONENTS COUNT",
            "TOTAL COMPONENTS ONBOARDED",
            "COMPONENTS PENDING TO BE ONBOARDED",
        ]].copy()
        sbom_tracker = sbom_tracker.rename(columns={"SBOM STATUS": "Status"})

        app_summary = combined_df.groupby(
            ["Business Segment", "Bussiness Lines", "APPLICATION"], as_index=False
        ).agg(
            Total_SBOMs=("SBOM", "count"),
            Completed_SBOMs=("SBOM Completed", "sum"),
            TOTAL_COMPONENTS_COUNT=("TOTAL COMPONENTS COUNT", "sum"),
            TOTAL_COMPONENTS_ONBOARDED=("TOTAL COMPONENTS ONBOARDED", "sum"),
            COMPONENTS_PENDING_TO_BE_ONBOARDED=("COMPONENTS PENDING TO BE ONBOARDED", "sum"),
        )
        total_sboms = pd.to_numeric(app_summary["Total_SBOMs"], errors="coerce").astype(float)
        completed_sboms = pd.to_numeric(app_summary["Completed_SBOMs"], errors="coerce").astype(float)
        app_summary["Completion %"] = (
            completed_sboms / total_sboms.replace(0, float("nan")) * 100
        ).round(1).fillna(0)

        app_status = app_summary.copy()
        app_status_flags = combined_df.copy()
        app_status_flags["_status_norm"] = app_status_flags["SBOM STATUS"].astype(str).str.strip().str.lower()
        app_status_flags = app_status_flags.groupby(
            ["Business Segment", "Bussiness Lines", "APPLICATION"], as_index=False
        ).agg(
            Has_Blocked=("_status_norm", lambda x: x.eq("blocked").any()),
            Has_On_Hold=("_status_norm", lambda x: x.isin(["on hold", "onhold"]).any()),
            Has_In_Progress=("_status_norm", lambda x: x.isin(["in progress", "inprogress"]).any()),
            InProgress_SBOMs=("_status_norm", lambda x: x.isin(["in progress", "inprogress"]).sum()),
            Blocked_SBOMs=("_status_norm", lambda x: x.eq("blocked").sum()),
            OnHold_SBOMs=("_status_norm", lambda x: x.isin(["on hold", "onhold"]).sum()),
        )

        app_status = app_status.merge(
            app_status_flags,
            on=["Business Segment", "Bussiness Lines", "APPLICATION"],
            how="left",
        )
        app_status[["Has_Blocked", "Has_On_Hold", "Has_In_Progress"]] = app_status[
            ["Has_Blocked", "Has_On_Hold", "Has_In_Progress"]
        ].fillna(False)
        app_status[["InProgress_SBOMs", "Blocked_SBOMs", "OnHold_SBOMs"]] = app_status[
            ["InProgress_SBOMs", "Blocked_SBOMs", "OnHold_SBOMs"]
        ].fillna(0).astype(int)

        def resolve_app_status(row):
            if row["Completed_SBOMs"] == row["Total_SBOMs"] and row["Total_SBOMs"] > 0:
                return "Completed"
            if row["Has_Blocked"]:
                return "Blocked"
            if row["Has_On_Hold"]:
                return "On Hold"
            if row["Has_In_Progress"] or row["Completed_SBOMs"] > 0:
                return "In Progress"
            return "In Progress"

        app_status["App Status"] = app_status.apply(resolve_app_status, axis=1)
        line_summary = app_status.groupby(
            ["Business Segment", "Bussiness Lines"], as_index=False
        ).agg(
            Total_Apps=("APPLICATION", "count"),
            Completed_Apps=("App Status", lambda x: (x == "Completed").sum()),
            In_Progress_Apps=("App Status", lambda x: (x == "In Progress").sum()),
            Blocked_Apps=("App Status", lambda x: (x == "Blocked").sum()),
            On_Hold_Apps=("App Status", lambda x: (x == "On Hold").sum()),
        )
        line_summary["Line Completed"] = line_summary["Completed_Apps"] == line_summary["Total_Apps"]
        
        # Define BL Status using same waterfall priority as App Status
        def resolve_bl_status(row):
            if row["Completed_Apps"] == row["Total_Apps"] and row["Total_Apps"] > 0:
                return "Completed"
            if row["Blocked_Apps"] > 0:
                return "Blocked"
            if row["On_Hold_Apps"] > 0:
                return "On Hold"
            if row["In_Progress_Apps"] > 0 or row["Completed_Apps"] > 0:
                return "In Progress"
            return "In Progress"
        
        line_summary["BL Status"] = line_summary.apply(resolve_bl_status, axis=1)
        total_apps = pd.to_numeric(line_summary["Total_Apps"], errors="coerce").astype(float)
        completed_apps = pd.to_numeric(line_summary["Completed_Apps"], errors="coerce").astype(float)
        line_summary["Completion %"] = (
            completed_apps / total_apps.replace(0, float("nan")) * 100
        ).round(1).fillna(0)

        seg_summary = line_summary.groupby("Business Segment", as_index=False).agg(
            Total_Lines=("Bussiness Lines", "count"),
            Completed_Lines=("Line Completed", "sum"),
            Blocked_Lines=("BL Status", lambda x: (x == "Blocked").sum()),
            On_Hold_Lines=("BL Status", lambda x: (x == "On Hold").sum()),
            In_Progress_Lines=("BL Status", lambda x: (x == "In Progress").sum()),
        )
        seg_summary["Segment Completed"] = seg_summary["Completed_Lines"] == seg_summary["Total_Lines"]
        
        # Define BS Status using same waterfall priority as App/BL Status
        def resolve_bs_status(row):
            if row["Completed_Lines"] == row["Total_Lines"] and row["Total_Lines"] > 0:
                return "Completed"
            if row["Blocked_Lines"] > 0:
                return "Blocked"
            if row["On_Hold_Lines"] > 0:
                return "On Hold"
            if row["In_Progress_Lines"] > 0 or row["Completed_Lines"] > 0:
                return "In Progress"
            return "In Progress"
        
        seg_summary["BS Status"] = seg_summary.apply(resolve_bs_status, axis=1)
        total_lines = pd.to_numeric(seg_summary["Total_Lines"], errors="coerce").astype(float)
        completed_lines = pd.to_numeric(seg_summary["Completed_Lines"], errors="coerce").astype(float)
        seg_summary["Completion %"] = (
            completed_lines / total_lines.replace(0, float("nan")) * 100
        ).round(1).fillna(0)

        total_components = pd.to_numeric(
            combined_df["TOTAL COMPONENTS COUNT"], errors="coerce"
        ).fillna(0).sum()
        total_onboarded_components = pd.to_numeric(
            combined_df["TOTAL COMPONENTS ONBOARDED"], errors="coerce"
        ).fillna(0).sum()

        waterfall = pd.DataFrame(
            {
                "Level": ["Components", "SBOMs", "Applications", "Business Lines", "Business Segments"],
                "Completed": [
                    float(total_onboarded_components),
                    int(combined_df["SBOM Completed"].sum()),
                    int((app_status["Completed_SBOMs"] == app_status["Total_SBOMs"]).sum()),
                    int(line_summary["Line Completed"].sum()),
                    int(seg_summary["Segment Completed"].sum()),
                ],
                "Total": [
                    float(total_components),
                    len(combined_df),
                    len(app_status),
                    len(line_summary),
                    len(seg_summary),
                ],
            }
        )
        waterfall_completed = pd.to_numeric(waterfall["Completed"], errors="coerce").astype(float)
        waterfall_total = pd.to_numeric(waterfall["Total"], errors="coerce").astype(float)
        waterfall["Completion %"] = (
            waterfall_completed / waterfall_total.replace(0, float("nan")) * 100
        ).round(1).fillna(0)

        comp_completion = combined_df.groupby("APPLICATION", as_index=False).agg(
            **{
                "Total Components": ("TOTAL COMPONENTS COUNT", "sum"),
                "Onboarded": ("TOTAL COMPONENTS ONBOARDED", "sum"),
                "Pending": ("COMPONENTS PENDING TO BE ONBOARDED", "sum"),
            }
        )
        comp_total = pd.to_numeric(comp_completion["Total Components"], errors="coerce").astype(float)
        comp_onboarded = pd.to_numeric(comp_completion["Onboarded"], errors="coerce").astype(float)
        comp_completion["Completion %"] = (
            comp_onboarded / comp_total.replace(0, float("nan")) * 100
        ).round(1).fillna(0)

        blocked_items = combined_df[combined_df["SBOM STATUS"].astype(str).str.lower() == "blocked"].copy()
        blocked_items = blocked_items[[
            "Business Segment",
            "Bussiness Lines",
            "APPLICATION",
            "SBOM",
            "SBOM STATUS",
            "TOTAL COMPONENTS COUNT",
            "COMPONENTS PENDING TO BE ONBOARDED",
            "JIRA TICKET",
        ]].rename(columns={"SBOM STATUS": "Status"})

        pending_by_segment = combined_df.groupby("Business Segment", as_index=False).agg(
            Total_SBOMs=("SBOM", "count"),
            Total_Components=("TOTAL COMPONENTS COUNT", "sum"),
            Onboarded=("TOTAL COMPONENTS ONBOARDED", "sum"),
            Pending=("COMPONENTS PENDING TO BE ONBOARDED", "sum"),
        )

        # Create unified hierarchy status sheet for slicers
        master_status = app_status[[
            "Business Segment",
            "Bussiness Lines",
            "APPLICATION",
            "App Status",
            "Total_SBOMs",
            "Completed_SBOMs",
            "InProgress_SBOMs",
            "Blocked_SBOMs",
            "OnHold_SBOMs",
            "Completion %",
            "TOTAL_COMPONENTS_COUNT",
            "TOTAL_COMPONENTS_ONBOARDED",
            "COMPONENTS_PENDING_TO_BE_ONBOARDED",
        ]].rename(columns={
            "Bussiness Lines": "Business Line",
            "APPLICATION": "Application",
            "App Status": "App Status",
            "Total_SBOMs": "Total SBOMs",
            "Completed_SBOMs": "Completed SBOMs",
            "InProgress_SBOMs": "In Progress SBOMs",
            "Blocked_SBOMs": "Blocked SBOMs",
            "OnHold_SBOMs": "On Hold SBOMs",
            "TOTAL_COMPONENTS_COUNT": "Total Components",
            "TOTAL_COMPONENTS_ONBOARDED": "Onboarded",
            "COMPONENTS_PENDING_TO_BE_ONBOARDED": "Pending",
        })
        master_status = master_status.merge(
            line_summary[["Business Segment", "Bussiness Lines", "BL Status"]].rename(
                columns={"Bussiness Lines": "Business Line"}
            ),
            on=["Business Segment", "Business Line"],
            how="left",
        )
        master_status = master_status.merge(
            seg_summary[["Business Segment", "BS Status"]],
            on=["Business Segment"],
            how="left",
        )
        master_status["BL Completed Key"] = master_status.apply(
            lambda row: row["Business Line"] if str(row["BL Status"]).strip().lower() == "completed" else "",
            axis=1,
        )
        master_status["App Completed Key"] = master_status.apply(
            lambda row: row["Application"] if str(row["App Status"]).strip().lower() == "completed" else "",
            axis=1,
        )
        # Reorder columns to match header list
        master_status = master_status[[
            "Business Segment", "Business Line", "Application", "App Status",
            "BL Status", "BS Status", "BL Completed Key", "App Completed Key",
            "Total SBOMs", "Completed SBOMs", "In Progress SBOMs", "Blocked SBOMs", "On Hold SBOMs",
            "Completion %", "Total Components", "Onboarded", "Pending"
        ]]

        return {
            "Master_Status": (
                master_status,
                ["Business Segment", "Business Line", "Application", "App Status",
                                    "BL Status", "BS Status", "BL Completed Key", "App Completed Key",
                 "Total SBOMs", "Completed SBOMs", "In Progress SBOMs", "Blocked SBOMs", "On Hold SBOMs",
                 "Completion %", "Total Components", "Onboarded", "Pending"],
            ),
            "SBOM_Tracker": (
                sbom_tracker,
                [
                    "Business Segment",
                    "Business Line",
                    "Application",
                    "SBOM",
                    "Status",
                    "Total Components",
                    "Onboarded",
                    "Pending",
                ],
            ),
            "App_Completion": (
                app_summary[[
                    "Business Segment",
                    "Bussiness Lines",
                    "APPLICATION",
                    "Completed_SBOMs",
                    "Total_SBOMs",
                    "Completion %",
                    "TOTAL_COMPONENTS_COUNT",
                    "TOTAL_COMPONENTS_ONBOARDED",
                    "COMPONENTS_PENDING_TO_BE_ONBOARDED",
                ]].rename(columns={
                    "Bussiness Lines": "Business Line",
                    "APPLICATION": "Application",
                    "TOTAL_COMPONENTS_COUNT": "Total Components",
                    "TOTAL_COMPONENTS_ONBOARDED": "Onboarded",
                    "COMPONENTS_PENDING_TO_BE_ONBOARDED": "Pending",
                }),
                ["Business Segment", "Business Line", "Application", "Completed SBOMs", "Total SBOMs", "Completion %", "Total Components", "Onboarded", "Pending"],
            ),
            "BL_Health": (
                line_summary[[
                    "Business Segment",
                    "Bussiness Lines",
                    "Total_Apps",
                    "Completed_Apps",
                    "In_Progress_Apps",
                    "Blocked_Apps",
                    "On_Hold_Apps",
                    "Completion %",
                ]].rename(columns={"Bussiness Lines": "Business Line"}),
                [
                    "Business Segment",
                    "Business Line",
                    "Total Apps",
                    "Completed Apps",
                    "In Progress Apps",
                    "Blocked Apps",
                    "On Hold Apps",
                    "Completion %",
                ],
            ),
            "Segment_Status": (
                seg_summary[["Business Segment", "Total_Lines", "Completed_Lines", "Completion %"]],
                ["Business Segment", "Total Lines", "Completed Lines", "Completion %"],
            ),
            "Completion_Waterfall": (
                waterfall,
                ["Level", "Completed", "Total", "Completion %"],
            ),
            "Component_Completion": (
                comp_completion,
                ["Application", "Total Components", "Onboarded", "Pending", "Completion %"],
            ),
            "Blocked_Items": (
                blocked_items.rename(columns={
                    "Bussiness Lines": "Business Line",
                    "APPLICATION": "Application",
                    "TOTAL COMPONENTS COUNT": "Total Components",
                    "COMPONENTS PENDING TO BE ONBOARDED": "Pending",
                    "JIRA TICKET": "JIRA Ticket",
                }),
                ["Business Segment", "Business Line", "Application", "SBOM", "Status", "Total Components", "Pending", "JIRA Ticket"],
            ),
            "Pending_Analysis": (
                pending_by_segment,
                ["Business Segment", "Total SBOMs", "Total Components", "Onboarded", "Pending"],
            ),
        }

    def _build_combined_data(self, input_path):
        excel_file = pd.ExcelFile(input_path)
        sheet_names = excel_file.sheet_names
        source_wb = load_workbook(input_path)

        excluded_sheets = {
            "dashboard",
            "pivot",
            "summary",
            "data",
            "masterstatus",
            "sbomtracker",
            "appcompletion",
            "blhealth",
            "segmentstatus",
            "completionwaterfall",
            "componentcompletion",
            "blockeditems",
            "pendinganalysis",
        }
        frames = []
        matched_data_sheets = []

        for sheet_name in sheet_names:
            if self._normalize(sheet_name) in excluded_sheets:
                continue

            segment = str(sheet_name).strip()
            if not segment:
                continue

            ws = source_wb[sheet_name]
            header_map = {}
            for col_idx in range(1, ws.max_column + 1):
                header_value = ws.cell(row=1, column=col_idx).value
                if header_value is None:
                    continue
                header_map[str(header_value).strip()] = col_idx

            missing = [col for col in self.REQUIRED_COLUMNS if col not in header_map]
            if missing:
                continue

            matched_data_sheets.append(sheet_name)

            row_records = []
            for row_idx in range(2, ws.max_row + 1):
                sbom_value = ws.cell(row=row_idx, column=header_map["SBOM"]).value
                if sbom_value is None or str(sbom_value).strip() == "":
                    continue

                record = {}
                for col_name in self.REQUIRED_COLUMNS:
                    raw_value = ws.cell(row=row_idx, column=header_map[col_name]).value
                    if col_name in self.NUMERIC_COLUMNS:
                        numeric_value = self._coerce_numeric_value(ws, raw_value)
                        record[col_name] = 0 if numeric_value is None else numeric_value
                    else:
                        record[col_name] = raw_value

                application_value = record.get("APPLICATION")
                record["APPLICATION"] = "" if application_value is None else str(application_value).strip()
                business_line_value = record.get("BUSSINESS LINES")
                record["BUSSINESS LINES"] = "" if business_line_value is None else str(business_line_value).strip()

                jira_cell = ws.cell(row=row_idx, column=header_map["JIRA TICKET"])
                jira_link = None
                if jira_cell.hyperlink is not None:
                    jira_link = jira_cell.hyperlink.target or jira_cell.hyperlink.location
                record["__JIRA_LINK"] = jira_link
                row_records.append(record)

            details = pd.DataFrame(
                row_records,
                columns=self.REQUIRED_COLUMNS + ["__JIRA_LINK"],
            )
            if details.empty:
                continue

            for col in self.NUMERIC_COLUMNS:
                details[col] = pd.to_numeric(details[col], errors="coerce").fillna(0)

            details.insert(0, "Business Segment", segment)
            details.rename(columns={"BUSSINESS LINES": "Bussiness Lines"}, inplace=True)
            frames.append(details)

        if not matched_data_sheets:
            raise ValueError(
                "No business segment sheets found with required columns: "
                + ", ".join(self.REQUIRED_COLUMNS)
            )

        if not frames:
            raise ValueError("No data rows found in business segment sheets")

        combined = pd.concat(frames, ignore_index=True)
        combined = combined.fillna("")
        return combined

    def run(self):
        try:
            self.progress_update.emit("🔄 Reading business segment sheets...")
            input_path = Path(self.input_file)

            if input_path.suffix.lower() == ".csv":
                raise ValueError("This tool requires an Excel workbook with business segment sheets")

            combined_df = self._build_combined_data(self.input_file)
            self.progress_update.emit(
                f"✅ Combined data ready: {combined_df.shape[0]} SBOM rows"
            )

            self.progress_update.emit("📊 Opening workbook...")
            workbook = load_workbook(self.input_file)
            existing_sheet_names = [ws.title for ws in workbook.worksheets]
            summary_frames = self._build_summary_frames(combined_df)
            summary_frames = {"Master_Status": summary_frames["Master_Status"]}
            obsolete_generated_sheets = [
                "SBOM_Tracker",
                "App_Completion",
                "BL_Health",
                "Segment_Status",
                "Completion_Waterfall",
                "Component_Completion",
                "Blocked_Items",
                "Pending_Analysis",
            ]

            def delete_sheet_case_insensitive(wb, target_name):
                target_norm = self._normalize_sheet_name(target_name)
                to_delete = [ws.title for ws in wb.worksheets if self._normalize_sheet_name(ws.title) == target_norm]
                for title in to_delete:
                    del wb[title]

            def write_df_openpyxl(ws, df, headers):
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                for row_idx, row_values in enumerate(df.values.tolist(), start=2):
                    for col_idx, value in enumerate(row_values, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border

                if ws.max_column > 0 and ws.max_row > 0:
                    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                ws.freeze_panes = "A2"

                for col_idx in range(1, ws.max_column + 1):
                    max_len = 0
                    for row_idx in range(1, ws.max_row + 1):
                        value = ws.cell(row=row_idx, column=col_idx).value
                        value_len = len(str(value)) if value is not None else 0
                        if value_len > max_len:
                            max_len = value_len
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

            self.progress_update.emit("📝 Writing data sheet...")
            delete_sheet_case_insensitive(workbook, "data")
            data_ws = workbook.create_sheet("data")

            data_df = combined_df[[col for col in combined_df.columns if col != "__JIRA_LINK"]].copy()
            data_headers = list(data_df.columns)
            write_df_openpyxl(data_ws, data_df, data_headers)

            jira_col_idx = data_headers.index("JIRA TICKET") + 1 if "JIRA TICKET" in data_headers else None
            if jira_col_idx is not None:
                for row_idx, jira_link in enumerate(combined_df["__JIRA_LINK"].tolist(), start=2):
                    if jira_link:
                        data_ws.cell(row=row_idx, column=jira_col_idx).hyperlink = str(jira_link)

            for obsolete_sheet in obsolete_generated_sheets:
                delete_sheet_case_insensitive(workbook, obsolete_sheet)

            self.progress_update.emit("📊 Creating master status sheet...")
            generated_sheet_names = list(summary_frames.keys())
            for sheet_name, (df, headers) in summary_frames.items():
                delete_sheet_case_insensitive(workbook, sheet_name)
                ws = workbook.create_sheet(sheet_name)
                write_df_openpyxl(ws, df.fillna(""), headers)

            # Order: Dashboard, data, Summary, Master, business segments, then generated sheets
            generated_norm = {self._normalize_sheet_name(name) for name in generated_sheet_names}
            priority_order = ["dashboard", "data", "summary", "master"]
            final_order = []

            all_sheet_names = [ws.title for ws in workbook.worksheets]

            for priority in priority_order:
                for name in all_sheet_names:
                    if self._normalize_sheet_name(name) == priority and name not in final_order:
                        final_order.append(name)
                        break

            for name in all_sheet_names:
                norm_name = self._normalize_sheet_name(name)
                if name in final_order:
                    continue
                if norm_name in generated_norm:
                    continue
                final_order.append(name)

            for name in generated_sheet_names:
                for current in all_sheet_names:
                    if self._normalize_sheet_name(current) == self._normalize_sheet_name(name) and current not in final_order:
                        final_order.append(current)
                        break

            for current in all_sheet_names:
                if current not in final_order:
                    final_order.append(current)

            workbook._sheets = [workbook[name] for name in final_order if name in workbook.sheetnames]

            self.progress_update.emit("📑 Created sheets: " + ", ".join([name for name in generated_sheet_names if name in workbook.sheetnames]))

            self.progress_update.emit("💾 Saving workbook...")
            workbook.save(self.input_file)

            self.progress_update.emit("✅ File updated successfully!")
            self.progress_complete.emit(True, f"✅ Excel file updated with Data + Master_Status sheets:\n{self.input_file}")
            
        except Exception as e:
            error_msg = f"❌ Error: {str(e)}\n\n{traceback.format_exc()}"
            self.progress_update.emit(error_msg)
            self.progress_complete.emit(False, error_msg)


class DashboardApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.worker = None
    
    def init_ui(self):
        """Initialize the user interface"""
        self.setWindowTitle("📊 Excel Dashboard Generator")
        self.setGeometry(100, 100, 800, 600)
        
        # Set application style
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f0f0f0;
            }
            QPushButton {
                background-color: #1F4E78;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #16395a;
            }
            QPushButton:pressed {
                background-color: #0f2437;
            }
            QLabel {
                font-size: 11px;
            }
            QTextEdit {
                background-color: white;
                border: 1px solid #cccccc;
                border-radius: 5px;
                padding: 5px;
            }
        """)
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title_label = QLabel("📊 Excel Dashboard Generator")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title_label.setFont(title_font)
        main_layout.addWidget(title_label)
        
        # Instructions
        info_label = QLabel("Upload an Excel file to generate a dashboard with statistics and analysis")
        info_label.setStyleSheet("color: #666666;")
        main_layout.addWidget(info_label)
        
        # File selection layout
        file_layout = QHBoxLayout()
        
        self.file_label = QLabel("No file selected")
        self.file_label.setStyleSheet("color: #1F4E78; font-weight: bold;")
        file_layout.addWidget(self.file_label)
        
        browse_btn = QPushButton("📁 Browse & Select Excel File")
        browse_btn.clicked.connect(self.select_file)
        file_layout.addWidget(browse_btn)
        
        main_layout.addLayout(file_layout)
        
        # Generate button
        generate_btn = QPushButton("🚀 Generate Dashboard")
        generate_font = QFont()
        generate_font.setPointSize(12)
        generate_font.setBold(True)
        generate_btn.setFont(generate_font)
        generate_btn.setMinimumHeight(50)
        generate_btn.clicked.connect(self.generate_dashboard)
        main_layout.addWidget(generate_btn)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #cccccc;
                border-radius: 5px;
                text-align: center;
                color: #1F4E78;
            }
            QProgressBar::chunk {
                background-color: #1F4E78;
            }
        """)
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)
        
        # Output text area
        output_label = QLabel("📝 Output Log:")
        output_label.setStyleSheet("font-weight: bold;")
        main_layout.addWidget(output_label)
        
        self.output_text = QTextEdit()
        self.output_text.setReadOnly(True)
        self.output_text.setMinimumHeight(200)
        main_layout.addWidget(self.output_text)
        
        # Footer buttons
        footer_layout = QHBoxLayout()
        
        clear_btn = QPushButton("🗑️ Clear Log")
        clear_btn.clicked.connect(self.clear_output)
        footer_layout.addWidget(clear_btn)
        
        footer_layout.addStretch()
        
        open_folder_btn = QPushButton("📂 Open Project Folder")
        open_folder_btn.clicked.connect(self.open_project_folder)
        footer_layout.addWidget(open_folder_btn)
        
        main_layout.addLayout(footer_layout)
        
        central_widget.setLayout(main_layout)
        
        self.selected_file = None
    
    def select_file(self):
        """Select an Excel file"""
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(
            self,
            "Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        
        if file_path:
            self.selected_file = file_path
            file_name = os.path.basename(file_path)
            self.file_label.setText(f"✅ Selected: {file_name}")
            self.log_output(f"Selected file: {file_path}\n")
    
    def generate_dashboard(self):
        """Generate dashboard from selected file"""
        if not self.selected_file:
            QMessageBox.warning(self, "No File Selected", "Please select an Excel file first!")
            return
        
        if not os.path.exists(self.selected_file):
            QMessageBox.critical(self, "File Not Found", "The selected file no longer exists!")
            return
        
        # Disable button and show progress
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.clear_output()
        self.log_output("Starting dashboard generation...\n")
        
        # Create worker thread
        self.worker = DashboardGenerator(self.selected_file)
        self.worker.progress_update.connect(self.on_progress_update)
        self.worker.progress_complete.connect(self.on_progress_complete)
        self.worker.start()
    
    def on_progress_update(self, message):
        """Handle progress updates"""
        self.log_output(message)
        self.progress_bar.setValue(min(self.progress_bar.value() + 20, 90))
    
    def on_progress_complete(self, success, message):
        """Handle completion"""
        self.progress_bar.setValue(100)
        self.log_output(f"\n{message}")
        
        if success:
            QMessageBox.information(self, "Success! ✅", message)
        else:
            QMessageBox.critical(self, "Error ❌", message)
    
    def log_output(self, message):
        """Add message to output log"""
        self.output_text.append(message)
        # Scroll to bottom
        self.output_text.verticalScrollBar().setValue(
            self.output_text.verticalScrollBar().maximum()
        )
    
    def clear_output(self):
        """Clear output log"""
        self.output_text.clear()
    
    def open_project_folder(self):
        """Open project folder"""
        project_path = os.path.dirname(os.path.abspath(__file__))
        os.startfile(project_path)


def main():
    app = QApplication(sys.argv)
    window = DashboardApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
